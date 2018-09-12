# -*- coding: utf-8 -*-
'''
EN:
This script converts xls, xlsx, csv files to mongodb collections.
You should install some packages before using this script. Use the
next command in your shell: "pip install name_of_packet". They are:
xlrd - to read Excel files;
openpyxl - to work with Excel files;
pandas - to cut down work previous libraries;
pymongo - to work with MongoDB.

The programm is:
1. Choose the way of processing - just one file or whole directory;
2. Depending on the way - choose the file with xls, xlsx or csv ext.
or go to 3d step;
3. Choose the directory and for each file do the next step;
4. If xls, xlsx - convert to csv, if csv - go directly to 5th step;
5. Convert to mongodb collection.

Python Version: 3.6.6.
Author: Ivan Lakshtovsky.

RU:
Данный скрипт преобразует файлы форматов xls, xlsx и csv в коллекции
mongodb. Перед использование нужно установить некоторые используемые пакеты.
Используйте команду "pip install имя_пакета". Итак, пакеты:
xlrd - для чтения файлов Excel;
openpyxl - ;
pandas - для упрощения работы двух предыдущих библиотек;
pymongo - для работы с MongoDB;

Ход программы:
1. Выбрать способ обработки - один файл или все из директории;
2. Если выбран первый способ - выбрать файл одного из трёх расширений,
перейти к шагу 4;
3. Если выбран второй способ - выбрать директорию. По всем файлам директории
прогнать шаг 4;
4. Если xls или xlsx, то перевести в csv, если csv, то сразу к шагу 5;
5. Перевести в коллекцию mongodb.

Версия Python: 3.6.6.
Автор: Иван Лакштовский.
'''
class FileWithExtFoundError(Exception):
    '''
    The user exception class for absence of needed formats
    '''
    pass

def excelToCSV(path):
    '''
    This function can convert xls/xlsx file to csv file by 3 different
    (and sometime similar) methods, using 3 dif. packages. The difference
    between packages has been showed at the descrtiption of this script.
    You can uncomment and use any approach. 
    '''
    #Firstly let's make name without extension.
    fname = path[:-4]
    if fname[-1] == '.': #If the .xlsx
        fname = fname[:-1]

    newNames = []
    '''
    #1st approach
    #There is some workaround because of excel representation numbers
    #like float.   
    wb = xlrd.open_workbook(path)
    shnames = wb.sheet_names()
    for shname in shnames:
        newName = fname + shname + '.csv' #Make a new file for each sheet
        with open(newName, 'w', newline = "") as newfile:
            sh = wb.sheet_by_name(shname)
            c = csv.writer(newfile)
            for r in range(sh.nrows):
                row_str = []
                for cellNum in range(0, len(sh.row_values(r))):
                    cell = sh.cell(r, cellNum) #We take one cell and check its' type.
                    cell_val = cell.value
                    if cell.ctype in (2,3) and int(cell_val) == cell_val:
                        cell_val = int(cell_val)
                    row_str.append(str(cell_val))
                c.writerow(row_str)
            newNames.append(newName)
    '''
    '''
    #2nd approach
    #openpyxl with this 'rows' help us avoid the previous problem.
    wb = openpyxl.load_workbook(path)
    shnames = wb.sheetnames
    for shname in shnames:
        newName = fname + shname + '.csv' #Make a new file for each sheet
        with open(newName, 'w', newline = "") as newfile:
            sh = wb[shname]
            c = csv.writer(newfile)
            for r in sh.rows:
                c.writerow([cell.value for cell in r])
            newNames.append(newName)
    '''
    #3d approach
    #Be careful with encoding of the path to file and encoding of the file.
    nf = pd.ExcelFile(path, dtype = str) #pd.read_excel(path, sheet_name = i, dtype = str)
    shnames = nf.sheet_names
    for shname in shnames:
        sh = nf.parse(shname)
        newName = fname + shname + '.csv'
        sh.to_csv(newName, index = False, encoding = 'cp1251') #encoding is optional, index destroys 'Unnamed' cell.
        newNames.append(newName)
        
    return newNames

def csvToMongo(allPaths):
    '''
    This procedure connect to mongoDB, convert csv to json objects and insert
    these objects to mongoDB collections.
    '''
    mConn = pm.MongoClient('localhost', 27017) #27017 - special MongoDb port.
    mDb = mConn['test'] #or mDb = mConn.test

    pattern = r'\w+'
    for onePath in allPaths:
        #'Collection' is another workaround to avoid invalid ns problem. The question is open. 
        colName = 'Collection' + re.findall(pattern, onePath)[-2]
        mColl = mDb[colName] #or mColl = mDb.collection if you know the name in advance.

        mColl.remove() ###

        data = pd.read_csv(open(onePath,'r')) #It's also some workaround to avoid OSError.
        #Another way is to use this commands: 1.import sys 2.sys._enablelegacywindowsfsencoding()
        #And third way is to use parameter engine = 'python'
        #Here is forth way is to use Mac OS or Unix-system, not windows.
        dataToRec = json.loads(data.to_json(orient = 'records')) #We convert it record by record.
        #Now, there is two way: 
        for oneRec in dataToRec:
            mColl.save(oneRec)
        #mColl.insert(dataToRec) #2d way

        #IT IS FOR DEBUGGING. PLEASE COMMENT OUT ON THE LARGE FILES!
        #print("Collections of " + onePath + " : ")
        #for obj in mColl.find():
        #    print(obj) 
    
def workWithDir():
    '''
    This procedure is to find all xls,xlsx,csv files in directory.
    You can comment out 'path = input' and uncomment
    'path = fd.askdirectory()'. Both variants are possible, but
    the askdirectory function made root empty window (can be repaired
    by new class. May be it will be done in next releases :) ).
    '''
    try:
        path = input("Enter the directory path, please, be careful with \\: ") #Get the path
        #path = fd.askdirectory()
        if path and not path[-1] == '\\':
            path += '\\'
            
        files = os.listdir(path) #Get the all files of this path
        files_xls = [path + file for file in files if file.endswith('.xls')
                     or file.endswith('.xlsx')] #Get the xls,xlsx file of files
        files_csv = [path + file for file in files if file.endswith('.csv')] #Get the csv files

        newFiles = []
        if files_csv:
            newFiles += files_csv

        if files_xls:
            for oneXls in files_xls:
                newFiles += excelToCSV(oneXls)
        elif not newFiles: #What should we do if nothing here?
            raise FileWithExtFoundError

        csvToMongo(newFiles)
        
    except FileNotFoundError: #except WindowsError: #for 2.7
        print("The directory path doesn't exist... Try another time...")
    except FileWithExtFoundError:
        print("The directory doesn't contain any xls,xlsx,csv files... Try another")

    
def workWithFile():
    '''
    This procedure is to find only one concrete file.
    '''
    try:
        #path = input("Enter the file name, please, be careful with \\: ") #Get the file name
        path = fd.askopenfilename()
            
        if not os.path.exists(path):
            raise FileNotFoundError #raise IOError
        
        if path.endswith('.csv'):
            newFiles = [path]
        elif path.endswith('.xls') or path.endswith('.xlsx'):
            newFiles = excelToCSV(path) #Make a csv files by xls/xlsx file.
        else:
            raise FileWithExtFoundError

        csvToMongo(newFiles)
        
    except FileNotFoundError: #except IOError: #for 2.7
        print("The file doesn't exist... Try another time...")
    except FileWithExtFoundError:
        print("The file not in xls/xlsx/csv formats... Try another")

if __name__ == '__main__':
    #Let's check all modules which will be used.
    try:
        import xlrd #Optional
        import openpyxl #Optional
        import pandas as pd #Optional
        import pymongo as pm
        import os
        import csv
        import json
        import re #It's just for some doings in csvToMongo func.
        from tkinter import filedialog as fd
            
    except ModuleNotFoundError as E: #except ImportError as E: #for 2.7
        print("It's seem that you haven't installed needed libraries yet.")
        print("Check the name: " + str(E))
        print('''Please use command shell and the next commands:
                \npip install xlrd
                \npip install openpyxl
                \npip install pandas"
                \npip install pymongo''') #Here is some problem in win system, install pymongo==3.4.0 
        quit()

    try:
        #Choose the way to solve the problem.
        way = '1'
        while (way == '1' or way == '2'):
            print("How would you like to handle with task?")
            print("Choose: ")
            print("1. - To process whole directory;")
            print("2. - To process the particular file;")
            print("other. - To exit.''')")
            way = input()
            if (way == '1'):
                workWithDir()
            elif (way == '2'):
                workWithFile()
    except Exception as E:
        print("Err... Some error has occured: ")
        print(E)

    print("Have a nice day!")
